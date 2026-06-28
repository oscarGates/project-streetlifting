#!/usr/bin/env python3
"""
Calculadora de ciclo 5/3/1 (2 ciclos = 8 semanas, con deload).

- Permite mezclar levantamientos de barra y de peso corporal (dips / pull-ups).
- Cada ejercicio se calcula y se MUESTRA en la unidad con la que ingresaste su
  RM: si lo das en kg el output va en kg (redondeo a 2.5 kg); si lo das en lbs,
  en lbs (redondeo a 5 lbs). Se puede mezclar unidades entre ejercicios.
- En dips/pull-ups el 1RM se ingresa como LASTRE anadido y el objetivo se
  expresa como "BW + X". Si un set de TRABAJO cae bajo el peso corporal se avisa
  y se calcula el lastre minimo necesario para evitarlo (en deload es normal que
  quede bajo el BW: se hace a peso corporal, sin alarma).
- Genera un Excel con 8 hojas (una por semana).
"""

from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    raise SystemExit("Falta la libreria openpyxl. Instala con: pip3 install openpyxl")


# --------------------------------------------------------------------------- #
# Constantes
# --------------------------------------------------------------------------- #
KG_TO_LBS = 2.20462
TM_FACTOR = 0.90            # Training Max = 90% del 1RM

# Incrementos por ciclo (en kg; se convierten si el ejercicio esta en lbs)
INCREMENT_UPPER_KG = 2.5
INCREMENT_LOWER_KG = 5.0

# Estructura de un ciclo de 4 semanas. Cada set: (porcentaje_TM, reps, es_amrap)
CYCLE = [
    {"nombre": "Semana 1 - 5/5/5+", "deload": False,
     "sets": [(0.65, 5, False), (0.75, 5, False), (0.85, 5, True)]},
    {"nombre": "Semana 2 - 3/3/3+", "deload": False,
     "sets": [(0.70, 3, False), (0.80, 3, False), (0.90, 3, True)]},
    {"nombre": "Semana 3 - 5/3/1+", "deload": False,
     "sets": [(0.75, 5, False), (0.85, 3, False), (0.95, 1, True)]},
    {"nombre": "Semana 4 - Deload", "deload": True,
     "sets": [(0.40, 5, False), (0.50, 5, False), (0.60, 5, False)]},
]

# Porcentaje de trabajo mas bajo (excluyendo deload). Define el lastre minimo.
MIN_WORK_PCT = min(pct for wk in CYCLE if not wk["deload"]
                   for (pct, _, _) in wk["sets"])


# --------------------------------------------------------------------------- #
# Utilidades
# --------------------------------------------------------------------------- #
def incremento_disco(unit):
    """Salto minimo de disco segun la unidad."""
    return 2.5 if unit == "kg" else 5.0


def redondear(x, unit):
    """Redondea al salto de disco de la unidad (2.5 kg o 5 lbs)."""
    inc = incremento_disco(unit)
    return round(x / inc) * inc


def convertir(valor, desde, hacia):
    if desde == hacia:
        return valor
    return valor * KG_TO_LBS if desde == "kg" else valor / KG_TO_LBS


def fmt(x):
    """Formatea un numero quitando el .0 innecesario."""
    return str(int(x)) if float(x).is_integer() else str(x)


def ask_float(prompt, minimo=None):
    while True:
        try:
            val = float(input(prompt).strip().replace(",", "."))
        except ValueError:
            print("  -> Ingresa un numero valido.")
            continue
        if minimo is not None and val <= minimo:
            print("  -> Debe ser mayor que %s." % fmt(minimo))
            continue
        return val


def ask_choice(prompt, opciones):
    """opciones: dict {tecla: descripcion}. Devuelve la tecla elegida."""
    opts_txt = " / ".join("[%s] %s" % (k, v) for k, v in opciones.items())
    while True:
        r = input("%s (%s): " % (prompt, opts_txt)).strip().lower()
        if r in opciones:
            return r
        print("  -> Opcion no valida.")


def ask_yes_no(prompt):
    return ask_choice(prompt, {"s": "si", "n": "no"}) == "s"


def ask_unit(prompt):
    return "kg" if ask_choice(prompt, {"k": "kg", "l": "lbs"}) == "k" else "lbs"


# --------------------------------------------------------------------------- #
# Captura de ejercicios
# --------------------------------------------------------------------------- #
def capturar_ejercicios():
    ejercicios = []
    print("=" * 60)
    print(" CALCULADORA DE CICLO 5/3/1  (2 ciclos / 8 semanas)")
    print("=" * 60)
    print("Vamos a registrar tus ejercicios uno por uno.\n")

    while True:
        nombre = input("Nombre del ejercicio (Enter para terminar): ").strip()
        if not nombre:
            if ejercicios:
                break
            print("  -> Agrega al menos un ejercicio.")
            continue

        es_bw = ask_yes_no(
            "  '%s' es de peso corporal (dips / pull-ups)?" % nombre)

        # Tren superior o inferior -> define el incremento entre ciclos.
        if es_bw:
            tren = "upper"   # dips y pull-ups son tren superior
        else:
            tren = "upper" if ask_choice(
                "  Tren del ejercicio", {"u": "superior", "l": "inferior"}
            ) == "u" else "lower"

        # 1RM. Para BW se ingresa como LASTRE anadido (estandar en calistenia).
        # La unidad del RM define la unidad de TODO el output del ejercicio.
        etiqueta_rm = "1RM (lastre anadido)" if es_bw else "1RM"
        unit = ask_unit("  Unidad del %s (define la unidad del output)"
                        % etiqueta_rm)
        rm = ask_float("  %s de '%s' en %s: " % (etiqueta_rm, nombre, unit),
                       minimo=0)

        # Peso corporal (solo bodyweight), convertido a la unidad del ejercicio.
        bw_disp = None
        if es_bw:
            unidad_bw = ask_unit("  Unidad del peso corporal")
            bw = ask_float("  Peso corporal en %s: " % unidad_bw, minimo=0)
            bw_disp = convertir(bw, unidad_bw, unit)

        # Carga total = cuerpo + lastre para BW; en barra es el 1RM tal cual.
        rm_total = (bw_disp + rm) if es_bw else rm
        tm = rm_total * TM_FACTOR
        ejercicios.append({
            "nombre": nombre,
            "es_bw": es_bw,
            "tren": tren,
            "unit": unit,
            "bw_disp": bw_disp,
            "tm_cycle1": tm,
        })
        print("  -> Registrado. TM (90%% del 1RM) = %s %s\n"
              % (fmt(redondear(tm, unit)), unit))

    return ejercicios


# --------------------------------------------------------------------------- #
# Calculo del programa
# --------------------------------------------------------------------------- #
def tm_para_ciclo(ej, ciclo):
    """TM (en la unidad del ejercicio) para el ciclo dado (0 o 1)."""
    tm = ej["tm_cycle1"]
    if ciclo == 1:
        inc_kg = INCREMENT_UPPER_KG if ej["tren"] == "upper" else INCREMENT_LOWER_KG
        tm += convertir(inc_kg, "kg", ej["unit"])
    return tm


def lastre_minimo(ej):
    """Lastre minimo (unidad del ejercicio) para que ningun set de TRABAJO
    caiga bajo el BW."""
    min_total = ej["bw_disp"] / (MIN_WORK_PCT * TM_FACTOR)
    return min_total - ej["bw_disp"]


def construir_semanas(ejercicios):
    """Devuelve lista de 8 semanas; cada una con filas por ejercicio/set."""
    semanas = []
    avisos_bw = {}  # nombre -> (lastre_min, unit)

    for ciclo in (0, 1):
        for wk in CYCLE:
            filas = []
            for ej in ejercicios:
                unit = ej["unit"]
                tm = tm_para_ciclo(ej, ciclo)
                for idx, (pct, reps, amrap) in enumerate(wk["sets"], 1):
                    objetivo = redondear(tm * pct, unit)
                    reps_txt = ("%d+" % reps) if amrap else str(reps)

                    if ej["es_bw"]:
                        bw = redondear(ej["bw_disp"], unit)
                        agregado = redondear(objetivo - ej["bw_disp"], unit)
                        if agregado >= 0:
                            detalle = "BW (%s %s) + %s %s" % (
                                fmt(bw), unit, fmt(agregado), unit)
                        elif wk["deload"]:
                            detalle = "Solo BW (sin lastre)"
                        else:
                            detalle = ("BW (%s %s) %s %s  <-- objetivo < BW!" % (
                                fmt(bw), unit, fmt(agregado), unit))
                            avisos_bw[ej["nombre"]] = (lastre_minimo(ej), unit)
                    else:
                        detalle = "%s %s en barra" % (fmt(objetivo), unit)

                    filas.append({
                        "ejercicio": ej["nombre"],
                        "set": idx,
                        "pct": int(round(pct * 100)),
                        "objetivo": "%s %s" % (fmt(objetivo), unit),
                        "detalle": detalle,
                        "reps": reps_txt,
                    })
            semanas.append({"nombre": wk["nombre"],
                            "ciclo": ciclo + 1,
                            "filas": filas})

    return semanas, avisos_bw


# --------------------------------------------------------------------------- #
# Exportar a Excel
# --------------------------------------------------------------------------- #
def exportar_excel(semanas, fname=None):
    """Genera el Excel. Si fname es None usa un nombre por defecto con fecha."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    title_font = Font(bold=True, size=13)
    amrap_fill = PatternFill("solid", fgColor="FCE4D6")
    center = Alignment(horizontal="center")
    cols = ["Ejercicio", "Set", "% TM", "Peso objetivo", "Detalle", "Reps"]

    for i, sem in enumerate(semanas, 1):
        ws = wb.create_sheet("Semana %d" % i)

        ws["A1"] = "%s  (Ciclo %d)" % (sem["nombre"], sem["ciclo"])
        ws["A1"].font = title_font
        ws.merge_cells("A1:F1")

        for c, titulo in enumerate(cols, 1):
            cell = ws.cell(row=2, column=c, value=titulo)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        r = 3
        for fila in sem["filas"]:
            es_amrap = fila["reps"].endswith("+")
            valores = [fila["ejercicio"], fila["set"], "%d%%" % fila["pct"],
                       fila["objetivo"], fila["detalle"], fila["reps"]]
            for c, v in enumerate(valores, 1):
                cell = ws.cell(row=r, column=c, value=v)
                if c in (2, 3, 6):
                    cell.alignment = center
                if es_amrap:
                    cell.fill = amrap_fill
            r += 1

        widths = [22, 6, 7, 16, 36, 8]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + c)].width = w
        ws.freeze_panes = "A3"

    if fname is None:
        fname = "ciclo_531_%s.xlsx" % date.today().isoformat()
    wb.save(fname)
    return fname


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ejercicios = capturar_ejercicios()
    semanas, avisos_bw = construir_semanas(ejercicios)

    # Resumen de Training Max por ciclo (cada uno en su unidad).
    print("\n" + "=" * 60)
    print(" RESUMEN DE TRAINING MAX")
    print("=" * 60)
    print("%-22s %-14s %-14s" % ("Ejercicio", "Ciclo 1", "Ciclo 2"))
    for ej in ejercicios:
        u = ej["unit"]
        print("%-22s %-14s %-14s" % (
            ej["nombre"],
            "%s %s" % (fmt(redondear(tm_para_ciclo(ej, 0), u)), u),
            "%s %s" % (fmt(redondear(tm_para_ciclo(ej, 1), u)), u)))

    # Avisos de peso corporal: solo si un set de trabajo cae bajo el BW.
    if avisos_bw:
        print("\n" + "!" * 60)
        print(" AVISO: objetivos de trabajo por debajo del peso corporal")
        print("!" * 60)
        for nombre, (lastre_min, u) in avisos_bw.items():
            print(" '%s': algun set de trabajo queda bajo tu BW." % nombre)
            print("   Lastre minimo en el 1RM para evitarlo: +%s %s"
                  % (fmt(redondear(lastre_min, u)), u))

    fname = exportar_excel(semanas)
    print("\nExcel generado: %s  (8 hojas, una por semana)\n" % fname)


if __name__ == "__main__":
    main()
